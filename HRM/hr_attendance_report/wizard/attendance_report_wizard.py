from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from datetime import datetime, date, timedelta
import io
import base64
from collections import defaultdict
import pytz

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    pass


class AttendanceReportWizard(models.TransientModel):
    _name = 'attendance.report.wizard'
    _description = 'Wizard Báo cáo Chấm công'

    # Filter fields
    filter_type = fields.Selection([
        ('date_range', 'Date Range'),
        ('month', 'By Month'),
    ], string='Filter Type', default='date_range', required=True)

    date_from = fields.Date(
        string='From Date',
        default=lambda self: fields.Date.today().replace(day=1),
        required=True
    )
    date_to = fields.Date(
        string='To Date',
        default=fields.Date.today(),
        required=True
    )

    month_year = fields.Date(
        string='Month/Year',
        default=fields.Date.today(),
        help='Select month for report'
    )

    employee_ids = fields.Many2many(
        'hr.employee',
        string='Employees',
        help='Leave empty for all employees'
    )

    department_ids = fields.Many2many(
        'hr.department',
        string='Departments',
        help='Leave empty for all departments'
    )

    report_type = fields.Selection([
        ('detailed', 'Detailed Report'),
        ('summary', 'Summary Report'),
        ('both', 'Both Types'),
    ], string='Report Type', default='detailed', required=True)

    include_discharge = fields.Boolean(
        string='Include Discharge Shifts',
        default=True,
        help='Include discharge shifts in report'
    )

    include_leave_data = fields.Boolean(
        string='Include Leave Data',
        default=True,
        help='Include leave data in report'
    )

    @api.onchange('filter_type')
    def _onchange_filter_type(self):
        """Thay đổi required fields theo loại lọc"""
        if self.filter_type == 'month':
            month_start = self.month_year.replace(day=1)
            if self.month_year.month == 12:
                month_end = self.month_year.replace(year=self.month_year.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = self.month_year.replace(month=self.month_year.month + 1, day=1) - timedelta(days=1)

            self.date_from = month_start
            self.date_to = month_end

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for record in self:
            if record.date_from > record.date_to:
                raise ValidationError(_('From date must be less than to date!'))

    def action_generate_report(self):
        """Tạo báo cáo Excel"""
        self.ensure_one()

        # Lấy dữ liệu attendance
        domain = self._get_attendance_domain()
        attendances = self.env['hr.attendance'].search(domain, order='employee_id, check_in')

        if not attendances:
            raise ValidationError(_('No attendance data found in selected period!'))

        # Lấy dữ liệu nghỉ phép nếu cần
        leave_data = {}
        leave_types = []
        if self.include_leave_data:
            leave_data, leave_types = self._get_leave_data()

        # Tạo file Excel
        output = io.BytesIO()
        workbook = Workbook()

        if self.report_type in ['detailed', 'both']:
            self._create_detailed_report(workbook, attendances, leave_data, leave_types)

        if self.report_type in ['summary', 'both']:
            self._create_summary_report(workbook, attendances, leave_data, leave_types)

        # Xóa sheet mặc định nếu có nhiều hơn 1 sheet
        if len(workbook.worksheets) > 1 and 'Báo cáo Chi tiết' not in workbook.sheetnames:
            workbook.remove(workbook.active)

        workbook.save(output)
        output.seek(0)

        # Tạo attachment
        filename = self._get_filename()
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _get_attendance_domain(self):
        """Tạo domain để lọc attendance - Fixed timezone conversion"""
        # Lấy timezone Vietnam
        vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
        utc_tz = pytz.UTC

        # Chuyển đổi date_from và date_to từ local date sang UTC datetime
        # date_from 00:00:00 Vietnam time -> UTC
        local_start = vietnam_tz.localize(datetime.combine(self.date_from, datetime.min.time()))
        utc_start = local_start.astimezone(utc_tz)

        # date_to 23:59:59 Vietnam time -> UTC
        local_end = vietnam_tz.localize(datetime.combine(self.date_to, datetime.max.time()))
        utc_end = local_end.astimezone(utc_tz)

        domain = [
            ('check_in', '>=', utc_start.replace(tzinfo=None)),  # Remove timezone info for Odoo
            ('check_in', '<=', utc_end.replace(tzinfo=None)),
            ('check_out', '!=', False),  # Chỉ lấy record đã check-out
        ]

        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))

        if self.department_ids:
            domain.append(('employee_id.department_id', 'in', self.department_ids.ids))

        if not self.include_discharge:
            domain.append(('is_discharge_shift', '=', False))

        return domain

    def _get_leave_data(self):
        """Lấy dữ liệu nghỉ phép trong khoảng thời gian"""
        leave_domain = [
            ('request_date_from', '<=', self.date_to),
            ('request_date_to', '>=', self.date_from),
            ('state', '=', 'validate'),
        ]

        if self.employee_ids:
            leave_domain.append(('employee_id', 'in', self.employee_ids.ids))

        if self.department_ids:
            leave_domain.append(('employee_id.department_id', 'in', self.department_ids.ids))

        leaves = self.env['hr.leave'].search(leave_domain)

        # Lấy danh sách các loại nghỉ phép
        leave_types = self.env['hr.leave.type'].search([])
        leave_type_dict = {lt.id: lt.name for lt in leave_types}

        # Gom dữ liệu theo nhân viên và ngày
        leave_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

        for leave in leaves:
            emp_id = leave.employee_id.id
            leave_type_name = leave_type_dict.get(leave.holiday_status_id.id, 'Khác')

            # Tính số ngày nghỉ trong khoảng thời gian báo cáo
            start_date = max(leave.request_date_from, self.date_from)
            end_date = min(leave.request_date_to, self.date_to)

            current_date = start_date
            while current_date <= end_date:
                # Tính số ngày nghỉ (8h = 1 ngày, 4h = 0.5 ngày)
                if leave.request_unit_half:
                    days = 0.5  # Nửa ngày
                elif leave.request_unit_hours:
                    days = 1.0  # Ngày nghỉ full
                else:
                    days = 1.0  # Mặc định ngày full

                leave_data[emp_id][current_date][leave_type_name] += days
                current_date += timedelta(days=1)

        return leave_data, list(leave_type_dict.values())

    def _create_detailed_report(self, workbook, attendances, leave_data=None, leave_types=None):
        """Tạo báo cáo chi tiết"""
        if 'Sheet' in workbook.sheetnames:
            ws = workbook.active
            ws.title = 'Báo cáo Chi tiết'
        else:
            ws = workbook.create_sheet('Báo cáo Chi tiết')

        # Header styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        title_col_count = 18 + len(leave_types or [])  # Giảm từ 19 xuống 18 do bỏ cột TC Nghỉ
        ws.merge_cells(f'A1:{get_column_letter(title_col_count)}1')
        ws[
            'A1'] = f'BÁO CÁO CHẤM CÔNG CHI TIẾT - {self.date_from.strftime("%d/%m/%Y")} đến {self.date_to.strftime("%d/%m/%Y")}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = [
            'STT', 'Mã NV', 'Tên nhân viên', 'Phòng ban', 'Ngày',
            'Giờ vào\n(Thực tế)', 'Giờ ra\n(Thực tế)', 'Giờ vào\n(Quy định)', 'Giờ ra\n(Quy định)',
            'Thời gian\nlàm việc', 'Tổng\nTăng ca', 'TC Sớm\n(Trước giờ làm việc)', 'TC Thường\n(Trước 18h)',
            'TC Tối\n(18h-21h)', 'TC Đêm\n(Sau 21h)', 'Trễ\n(phút)', 'Sớm\n(phút)', 'Xả ca'
        ]

        # Thêm cột nghỉ phép
        if leave_types:
            for leave_type in sorted(leave_types):
                headers.append(f'NP {leave_type}\n(ngày)')

        # Set headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Adjust column widths
        column_widths = [5, 12, 20, 15, 12, 12, 12, 12, 12, 10, 10, 8, 10, 8, 8, 8, 8,
                         8]  # Bỏ 1 width do xóa cột TC Nghỉ
        # Thêm width cho cột nghỉ phép
        if leave_types:
            column_widths.extend([10] * len(leave_types))

        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
        row = 4

        for idx, attendance in enumerate(attendances, 1):
            # Convert times to Vietnam timezone for display
            check_in_vn = attendance.check_in.replace(tzinfo=pytz.UTC).astimezone(vietnam_tz)
            check_out_vn = attendance.check_out.replace(tzinfo=pytz.UTC).astimezone(vietnam_tz)

            scheduled_in_vn = scheduled_out_vn = ''
            if attendance.scheduled_check_in:
                scheduled_in_vn = attendance.scheduled_check_in.replace(tzinfo=pytz.UTC).astimezone(
                    vietnam_tz).strftime('%H:%M')
            if attendance.scheduled_check_out:
                scheduled_out_vn = attendance.scheduled_check_out.replace(tzinfo=pytz.UTC).astimezone(
                    vietnam_tz).strftime('%H:%M')

            data = [
                idx,  # STT
                attendance.employee_id.employee_code or '',  # Mã NV
                attendance.employee_id.name,  # Tên NV
                attendance.employee_id.department_id.name or '',  # Phòng ban
                check_in_vn.strftime('%d/%m/%Y'),  # Ngày
                check_in_vn.strftime('%H:%M'),  # Giờ vào thực tế
                check_out_vn.strftime('%H:%M'),  # Giờ ra thực tế
                scheduled_in_vn,  # Giờ vào quy định
                scheduled_out_vn,  # Giờ ra quy định
                f"{attendance.worked_hours:.1f}",  # Thời gian làm việc
                f"{attendance.overtime_hours:.1f}",  # Tổng tăng ca
                f"{attendance.overtime_early:.1f}",  # TC sớm
                f"{attendance.overtime_regular:.1f}",  # TC thường
                f"{attendance.overtime_evening:.1f}",  # TC tối
                f"{attendance.overtime_night:.1f}",  # TC đêm
                attendance.late_minutes,  # Đi trễ
                attendance.early_minutes,  # Về sớm
                'Có' if attendance.is_discharge_shift else ''  # Xả ca
            ]

            # Thêm dữ liệu nghỉ phép
            if leave_types and leave_data:
                emp_id = attendance.employee_id.id
                work_date = check_in_vn.date()
                for leave_type in sorted(leave_types):
                    leave_days = leave_data.get(emp_id, {}).get(work_date, {}).get(leave_type, 0)
                    if leave_days > 0:
                        if leave_days == 1.0:
                            data.append('1')
                        elif leave_days == 0.5:
                            data.append('0.5')
                        else:
                            data.append(f'{leave_days:.1f}')
                    else:
                        data.append('')

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border

                # Alignment based on column type
                base_center_cols = [1, 5, 6, 7, 8, 16, 17, 18]  # Điều chỉnh từ [1, 5, 6, 7, 8, 17, 18, 19]
                leave_start_col = 19  # Giảm từ 20 xuống 19
                if col in base_center_cols or (leave_types and col >= leave_start_col):
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

                # Color coding
                if col == 16 and attendance.late_minutes > 0:  # Late (điều chỉnh từ 17 xuống 16)
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif col == 17 and attendance.early_minutes > 0:  # Early (điều chỉnh từ 18 xuống 17)
                    cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                elif col == 18 and attendance.is_discharge_shift:  # Discharge (điều chỉnh từ 19 xuống 18)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                elif leave_types and col >= leave_start_col and value and value != '':  # Leave
                    cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

            row += 1

    def _get_employee_timezone(self, employee):
        """
        Get employee timezone for proper date/time calculations
        """
        # Try to get timezone from employee resource calendar
        if employee.resource_calendar_id and employee.resource_calendar_id.tz:
            try:
                return pytz.timezone(employee.resource_calendar_id.tz)
            except:
                pass

        # Try to get timezone from company
        if employee.company_id and hasattr(employee.company_id, 'resource_calendar_id'):
            if employee.company_id.resource_calendar_id and employee.company_id.resource_calendar_id.tz:
                try:
                    return pytz.timezone(employee.company_id.resource_calendar_id.tz)
                except:
                    pass

        # Fallback to Vietnam timezone
        return pytz.timezone('Asia/Ho_Chi_Minh')

    def _calculate_work_days_for_attendance(self, attendance, leave_data=None):
        """
        Tính số ngày công cho một attendance record theo logic của payslip

        Args:
            attendance: hr.attendance record
            leave_data: dict chứa dữ liệu nghỉ phép theo ngày

        Returns:
            float: Số ngày công (0.0, 0.5, hoặc 1.0)
        """
        if not attendance.check_out or attendance.worked_hours <= 0:
            return 0.0

        employee = attendance.employee_id
        employee_tz = self._get_employee_timezone(employee)

        # Convert UTC check_in time to employee local time for date calculation
        check_in_local = attendance.check_in.replace(tzinfo=pytz.UTC).astimezone(employee_tz)
        attendance_date = check_in_local.date()

        # Get work calendar
        work_calendar = employee.resource_calendar_id
        if not work_calendar:
            # Fallback to standard calculation if no calendar
            return self._calculate_work_days_standard(attendance.worked_hours)

        # QUAN TRỌNG: Lấy scheduled_hours cho NGUYÊN NGÀY (00:00 - 23:59)
        # Tạo datetime cho nguyên ngày ở múi giờ UTC
        day_start_utc = datetime.combine(attendance_date, datetime.min.time()).replace(tzinfo=pytz.UTC)
        day_end_utc = datetime.combine(attendance_date, datetime.max.time()).replace(tzinfo=pytz.UTC)

        # Sử dụng get_work_hours_count để lấy tổng giờ làm việc theo lịch cho nguyên ngày
        try:
            scheduled_hours = work_calendar.get_work_hours_count(
                day_start_utc,
                day_end_utc,
                compute_leaves=True
            )
        except:
            # Fallback nếu method get_work_hours_count fail
            # Thử cách manual tính từ intervals
            day_start_local = employee_tz.localize(datetime.combine(attendance_date, datetime.min.time()))
            day_end_local = employee_tz.localize(datetime.combine(attendance_date, datetime.max.time()))

            try:
                intervals = work_calendar._work_intervals_batch(
                    day_start_local,
                    day_end_local,
                    employee.resource_id
                )[employee.resource_id.id]

                scheduled_hours = 0.0
                work_intervals = [(start, stop) for start, stop, meta in intervals]

                for start, stop in work_intervals:
                    scheduled_hours += (stop - start).total_seconds() / 3600.0

            except:
                # Fallback cuối cùng
                return self._calculate_work_days_standard(attendance.worked_hours)

        # Nếu không có giờ làm việc theo lịch thì đây là ngày nghỉ
        if scheduled_hours <= 0:
            return 0.0  # Non-working day

        # Determine if it's full day or half day based on worked hours vs scheduled hours
        worked_day_ratio = min((attendance.worked_hours - attendance.overtime_hours)/ scheduled_hours, 1.0)

        # Apply threshold logic to determine full/half day (same as payslip)
        if worked_day_ratio >= 0.70:  # 70% threshold for full day
            worked_day_amount = 1.0
        elif worked_day_ratio >= 0.375:  # 37.5% threshold for half day
            worked_day_amount = 0.5
        else:
            worked_day_amount = 0.0

        return worked_day_amount

    def _calculate_work_days_standard(self, worked_hours, standard_hours_per_day=8.0):
        """
        Fallback method để tính ngày công khi không có resource calendar

        Args:
            worked_hours (float): Số giờ làm việc thực tế
            standard_hours_per_day (float): Số giờ tiêu chuẩn của 1 ngày công

        Returns:
            float: Số ngày công
        """
        if worked_hours <= 0:
            return 0.0

        # Tính tỷ lệ giờ làm việc so với giờ tiêu chuẩn
        worked_day_ratio = worked_hours / standard_hours_per_day

        # Áp dụng ngưỡng giống như trong payslip
        if worked_day_ratio >= 0.75:  # 75% threshold for full day
            return 1.0
        elif worked_day_ratio >= 0.375:  # 37.5% threshold for half day
            return 0.5
        else:
            return 0.0

    def _create_summary_report(self, workbook, attendances, leave_data=None, leave_types=None):
        """Tạo báo cáo tổng hợp - Fixed work days calculation using payslip logic"""
        ws = workbook.create_sheet('Báo cáo Tổng hợp')

        # Gom dữ liệu theo nhân viên
        employee_data = defaultdict(lambda: {
            'name': '',
            'employee_id': '',
            'department': '',
            'work_days': 0.0,  # Changed to float to handle half days
            'total_worked_hours': 0.0,
            'total_overtime': 0.0,
            'overtime_early': 0.0,
            'overtime_regular': 0.0,
            'overtime_evening': 0.0,
            'overtime_night': 0.0,
            'overtime_holiday': 0.0,
            'total_late_minutes': 0,
            'total_early_minutes': 0,
            'discharge_days': 0,
            'late_days': 0,
            'early_days': 0,
            'leave_days': defaultdict(float),
        })
        count = 1
        for attendance in attendances:
            emp_id = attendance.employee_id.id
            data = employee_data[emp_id]

            # Basic info
            data['name'] = attendance.employee_id.name
            data['employee_id'] = attendance.employee_id.employee_code or ''
            data['department'] = attendance.employee_id.department_id.name or ''

            # Calculate work days using payslip logic
            work_days_for_this_attendance = self._calculate_work_days_for_attendance(attendance, leave_data)
            count += 1

            # Accumulate data
            data['work_days'] += work_days_for_this_attendance  # Use payslip logic
            data['total_worked_hours'] += attendance.worked_hours
            data['total_overtime'] += attendance.overtime_hours
            data['overtime_early'] += attendance.overtime_early
            data['overtime_regular'] += attendance.overtime_regular
            data['overtime_evening'] += attendance.overtime_evening
            data['overtime_night'] += attendance.overtime_night
            data['overtime_holiday'] += attendance.overtime_holiday
            data['total_late_minutes'] += attendance.late_minutes
            data['total_early_minutes'] += attendance.early_minutes

            if attendance.is_discharge_shift:
                data['discharge_days'] += 1
            if attendance.is_late:
                data['late_days'] += 1
            if attendance.is_early:
                data['early_days'] += 1

        # Gom dữ liệu nghỉ phép
        if leave_data and leave_types:
            for emp_id, emp_leaves in leave_data.items():
                if emp_id in employee_data:
                    for date_leaves in emp_leaves.values():
                        for leave_type, days in date_leaves.items():
                            employee_data[emp_id]['leave_days'][leave_type] += days

        # Header styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        title_col_count = 17 + len(leave_types or [])  # Giảm từ 18 xuống 17 do bỏ 1 cột TC Nghỉ
        ws.merge_cells(f'A1:{get_column_letter(title_col_count)}1')
        ws[
            'A1'] = f'BÁO CÁO CHẤM CÔNG TỔNG HỢP - {self.date_from.strftime("%d/%m/%Y")} đến {self.date_to.strftime("%d/%m/%Y")}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = [
            'STT', 'Mã NV', 'Tên nhân viên', 'Phòng ban',
            'Tổng\nNgày công', 'Tổng giờ\nlàm việc', 'Tổng\nTăng ca',
            'TC Sớm\n(Trước giờ làm việc)', 'TC Thường\n(Trước 18h)', 'TC Tối\n(18h-21h)', 'TC Đêm\n(Sau 21h)',
            'Tổng phút\nđi trễ', 'Ngày\nđi trễ', 'Tổng phút\nvề sớm', 'Ngày\nvề sớm',
            'Tổng phút\nđi trễ-về sớm',  # Cột mới thêm
            'Số ca xả'
        ]

        # Thêm cột nghỉ phép
        if leave_types:
            for leave_type in sorted(leave_types):
                headers.append(f'Tổng NP\n{leave_type}\n(ngày)')

        # Set headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Adjust column widths
        column_widths = [5, 12, 20, 15, 10, 12, 10, 8, 10, 8, 8, 10, 8, 10, 8, 12, 8]  # Bỏ 1 width do xóa cột TC Nghỉ
        # Thêm width cho cột nghỉ phép
        if leave_types:
            column_widths.extend([12] * len(leave_types))

        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        row = 4
        for idx, (emp_id, data) in enumerate(sorted(employee_data.items(), key=lambda x: x[1]['name']), 1):
            # Tính tổng phút đi trễ + về sớm
            total_late_early_minutes = data['total_late_minutes'] + data['total_early_minutes']

            # Format work days display - using payslip logic formatting
            work_days_display = data['work_days']
            if work_days_display == int(work_days_display):
                work_days_display = int(work_days_display)
            else:
                work_days_display = f"{work_days_display:.1f}"

            summary_data = [
                idx,  # STT
                data['employee_id'],  # Mã NV
                data['name'],  # Tên NV
                data['department'],  # Phòng ban
                work_days_display,  # Tổng ngày công - Using payslip logic
                f"{data['total_worked_hours']:.1f}",  # Tổng giờ làm việc
                f"{data['total_overtime']:.1f}",  # Tổng tăng ca
                f"{data['overtime_early']:.1f}",  # TC sớm
                f"{data['overtime_regular']:.1f}",  # TC thường
                f"{data['overtime_evening']:.1f}",  # TC tối
                f"{data['overtime_night']:.1f}",  # TC đêm
                data['total_late_minutes'],  # Tổng phút đi trễ
                data['late_days'],  # Số ngày đi trễ
                data['total_early_minutes'],  # Tổng phút về sớm
                data['early_days'],  # Số ngày về sớm
                total_late_early_minutes,  # Tổng phút đi trễ-về sớm (cột mới)
                data['discharge_days'],  # Số ca xả
            ]

            # Thêm dữ liệu nghỉ phép
            if leave_types:
                for leave_type in sorted(leave_types):
                    leave_days = data['leave_days'].get(leave_type, 0)
                    if leave_days > 0:
                        if leave_days == int(leave_days):  # Số nguyên
                            summary_data.append(f'{int(leave_days)}')
                        else:  # Số thập phân
                            summary_data.append(f'{leave_days:.1f}')
                    else:
                        summary_data.append('')

            for col, value in enumerate(summary_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border

                # Alignment
                base_center_cols = [1, 5, 12, 13, 14, 15, 16, 17]  # Điều chỉnh lại số cột do bỏ cột TC Nghỉ
                leave_start_col = 18  # Giảm từ 19 xuống 18 do bỏ 1 cột
                if col in base_center_cols or (leave_types and col >= leave_start_col):
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

                # Color coding for issues
                if col == 12 and data['total_late_minutes'] > 0:  # Late minutes (điều chỉnh từ 13 xuống 12)
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif col == 14 and data['total_early_minutes'] > 0:  # Early minutes (điều chỉnh từ 15 xuống 14)
                    cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                elif col == 16 and total_late_early_minutes > 0:  # Tổng phút đi trễ-về sớm (điều chỉnh từ 17 xuống 16)
                    cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')  # Màu cam nhạt
                elif col == 17 and data['discharge_days'] > 0:  # Discharge days (điều chỉnh từ 18 xuống 17)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                elif leave_types and col >= leave_start_col and value and value != '':  # Leave
                    cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

            row += 1

    def _get_filename(self):
        """Generate filename"""
        date_str = f"{self.date_from.strftime('%d%m%Y')}-{self.date_to.strftime('%d%m%Y')}"

        if self.report_type == 'detailed':
            type_str = 'ChiTiet'
        elif self.report_type == 'summary':
            type_str = 'TongHop'
        else:
            type_str = 'DayDu'

        return f'BaoCaoChamCong_{type_str}_{date_str}.xlsx'

    def action_preview_data(self):
        """Preview data before generating report"""
        domain = self._get_attendance_domain()

        return {
            'name': 'Xem trước dữ liệu báo cáo',
            'type': 'ir.actions.act_window',
            'res_model': 'hr.attendance',
            'view_mode': 'tree,form',
            'domain': domain,
            'context': {
                'search_default_groupby_employee_code': 1,
            },
            'target': 'new',
        }